import pandas
import glob #is used to make a list of file paths
from fpdf import FPDF
import openpyxl #it has to be imported to read xlsx files
import pathlib


filepaths = glob.glob(r"excelFiles\*.xlsx")

for filepath in filepaths:
    pdf = FPDF(orientation="P",unit="mm",format="A4")
    filename= pathlib.Path(filepath).stem #shows exact filename
    invoice_number = filename.split("-")[0]
    date = filename.split("-")[1]
    pdf.add_page()
    pdf.set_font(family="times",style="B",size=15)
    pdf.cell(w=50,h=8,txt=f"Invoice number: {invoice_number}",ln=1)
    pdf.cell(w=50,h=8,txt=f"date: {date}",ln=1)

    data = pandas.read_excel(filepath, sheet_name="Sheet 1")
    header =  list(data.columns)
    header = [item.replace("_"," ") for item in header]
    pdf.set_font(family="times", size=10,style="B")
    pdf.cell(w=30, h=8, txt=header[0], ln=0, border=1)
    pdf.cell(w=70, h=8, txt=header[1], ln=0, border=1)
    pdf.cell(w=30, h=8, txt=header[2], ln=0, border=1)
    pdf.cell(w=30, h=8, txt=header[3], ln=0, border=1)
    pdf.cell(w=30, h=8, txt=header[4], ln=1, border=1)

    total = 0
    for index, rows in data.iterrows():
        pdf.set_font(family="times",size=10)
        pdf.set_text_color(100,100,100)
        pdf.cell(w=30,h=8,txt=str(rows["product_id"]),ln=0,border=1)
        pdf.cell(w=70, h=8, txt=str(rows["product_name"]),ln=0,border=1)
        pdf.cell(w=30, h=8, txt=str(rows["amount_purchased"]),ln=0,border=1)
        pdf.cell(w=30, h=8, txt=str(rows["price_per_unit"]),ln=0,border=1)
        pdf.cell(w=30, h=8, txt=str(rows["total_price"]),ln=1,border=1)
        total+= int(rows["total_price"])

    pdf.set_font(family="times", size=10,style="B")
    pdf.set_text_color(0,0,0)
    pdf.cell(w=30, h=8, txt="", ln=0, border=1)
    pdf.cell(w=70, h=8, txt="", ln=0, border=1)
    pdf.cell(w=30, h=8, txt="", ln=0, border=1)
    pdf.cell(w=30, h=8, txt="", ln=0, border=1)
    pdf.cell(w=30, h=8, txt=str(total), ln=1, border=1)

    pdf.ln(10)

    pdf.set_font(family="times", style="B", size=13)
    pdf.cell(w=50, h=8, txt=f"Your total is: {total}", ln=1)


    pdf.output(fr"invoicePDFS\{filename}.pdf")